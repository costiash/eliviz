#!/usr/bin/env python3
"""
parse_input.py — detect and normalize data files of many formats into the
model consumed by assets/template.html.

Supported inputs (auto-detected by extension, then content sniffing):
  - Tabular:  .csv .tsv .jsonl .ndjson .xlsx .xlsm .sqlite .sqlite3 .db
              (+ JSON arrays of flat records are promoted to tables)
  - JSON:     .json — arrays of records become tables; everything else
              becomes the structured tree explorer ("generic")
  - Text:     .md .markdown (rendered reading view) and plain .txt
  - Logs:     .log, or any text file that looks like a log (levels/timestamps)

Usage:
  python parse_input.py input1 [input2 ...] [-o normalized.json]
                        [--title T] [--subtitle S] [--accent HEX]
                        [--max-rows N] [--max-text-chars N] [--max-log-lines N]

See references/data-model.md for the exact normalized schema.
"""

import argparse
import csv
import io
import json
import math
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone

MAX_GENERIC_BYTES = 2_000_000
GENERIC_ARRAY_CAP = 200
MAX_COLS = 60
CELL_STR_CAP = 300
DEFAULT_MAX_ROWS = 2000
DEFAULT_MAX_TEXT_CHARS = 400_000
DEFAULT_MAX_LOG_LINES = 5000
HIST_BINS = 20
SHEET_CAP = 8
SQL_TABLE_CAP = 8
SQL_ROW_SCAN_CAP = 100_000

TABULAR_EXTS = {"csv", "tsv", "jsonl", "ndjson", "xlsx", "xlsm",
                "sqlite", "sqlite3", "db"}


# ---------------------------------------------------------------- helpers

def clean(v):
    """Treat the literal strings 'None'/'null'/'' as missing."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() in ("", "None", "null", "NULL", "NaN", "nan"):
        return None
    return v


def parse_dt(s):
    s = clean(s)
    if isinstance(s, datetime):
        return s
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%d-%m-%Y", "%m/%d/%Y",
                "%d/%m/%Y", "%b %d %Y", "%d %b %Y", "%Y-%m"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def word_count(text):
    return len(re.findall(r"\S+", text)) if text else 0


def cap_text(text, limit):
    if text is None:
        return "", False
    if limit and len(text) > limit:
        extra = len(text) - limit
        return text[:limit] + f"\n… [truncated {extra:,} more chars]", True
    return text, False


def fmt_num(x):
    """Compact label formatting for histogram bin edges."""
    if x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x))):
        return "?"
    ax = abs(x)
    if ax >= 1_000_000:
        return f"{x/1_000_000:.1f}M".replace(".0M", "M")
    if ax >= 10_000:
        return f"{x/1000:.0f}k"
    if ax >= 1000:
        return f"{x/1000:.1f}k".replace(".0k", "k")
    if isinstance(x, float) and not x.is_integer():
        return f"{x:.2f}".rstrip("0").rstrip(".")
    return str(int(x))


# ---------------------------------------------------------------- type inference

_NUM_RE = re.compile(r"^-?(\d{1,3}(,\d{3})+|\d+)(\.\d+)?([eE][+-]?\d+)?$")
_BOOL_TRUE = {"true", "yes"}
_BOOL_FALSE = {"false", "no"}


def try_number(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and math.isnan(v)) else float(v)
    if isinstance(v, str) and _NUM_RE.match(v.strip()):
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def try_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        low = v.strip().lower()
        if low in _BOOL_TRUE:
            return True
        if low in _BOOL_FALSE:
            return False
    return None


def infer_column_type(values):
    """values: non-null sample. Returns 'number'|'boolean'|'date'|'string'."""
    if not values:
        return "string"
    n = len(values)
    threshold = max(1, int(n * 0.9))
    if sum(1 for v in values if try_bool(v) is not None) >= threshold:
        return "boolean"
    if sum(1 for v in values if try_number(v) is not None) >= threshold:
        return "number"
    # dates are expensive to parse — check a subsample first
    sub = values[:80]
    if sum(1 for v in sub if parse_dt(v) is not None) >= max(1, int(len(sub) * 0.9)):
        if sum(1 for v in values if parse_dt(v) is not None) >= threshold:
            return "date"
    return "string"


# ---------------------------------------------------------------- table dataset

def cell_repr(v):
    """A display-safe cell value for embedding in the page."""
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, (dict, list)):
        s = json.dumps(v, ensure_ascii=False)
    else:
        s = str(v)
    return s if len(s) <= CELL_STR_CAP else s[:CELL_STR_CAP] + "…"


def profile_column(name, values):
    """values: full list including Nones. Returns the column profile dict."""
    non_null = [v for v in values if clean(v) is not None]
    ctype = infer_column_type(non_null)
    prof = {
        "name": name,
        "type": ctype,
        "nulls": len(values) - len(non_null),
        "distinct": 0,
    }
    if ctype == "number":
        nums = [n for n in (try_number(v) for v in non_null) if n is not None]
        prof["distinct"] = len(set(nums))
        if nums:
            nums.sort()
            lo, hi = nums[0], nums[-1]
            mid = len(nums) // 2
            median = nums[mid] if len(nums) % 2 else (nums[mid - 1] + nums[mid]) / 2
            prof.update({
                "min": lo, "max": hi,
                "sum": round(sum(nums), 4),
                "mean": round(sum(nums) / len(nums), 4),
                "median": round(median, 4),
            })
            if hi > lo:
                width = (hi - lo) / HIST_BINS
                counts = [0] * HIST_BINS
                for x in nums:
                    idx = min(int((x - lo) / width), HIST_BINS - 1)
                    counts[idx] += 1
                prof["histogram"] = [
                    {"label": f"{fmt_num(lo + i * width)}–{fmt_num(lo + (i + 1) * width)}",
                     "count": c}
                    for i, c in enumerate(counts)
                ]
            else:
                prof["histogram"] = [{"label": fmt_num(lo), "count": len(nums)}]
    elif ctype == "boolean":
        bools = [b for b in (try_bool(v) for v in non_null) if b is not None]
        t = sum(1 for b in bools if b)
        prof["distinct"] = len(set(bools))
        prof["true_count"] = t
        prof["false_count"] = len(bools) - t
        prof["true_pct"] = round(100 * t / len(bools), 1) if bools else 0
    elif ctype == "date":
        dts = [d for d in (parse_dt(v) for v in non_null) if d is not None]
        prof["distinct"] = len({d.isoformat() for d in dts})
        if dts:
            naive = [d.replace(tzinfo=None) for d in dts]
            prof["min"] = min(naive).isoformat()
            prof["max"] = max(naive).isoformat()
    else:
        strs = [str(v) for v in non_null]
        counter = Counter(strs)
        prof["distinct"] = len(counter)
        prof["top_values"] = [
            {"value": (v if len(v) <= 80 else v[:80] + "…"), "count": c}
            for v, c in counter.most_common(10)
        ]
        if strs:
            prof["avg_len"] = round(sum(len(s) for s in strs) / len(strs), 1)
    return prof


def build_timeseries(columns, col_values):
    """Pick the best date column and bucket row counts over time."""
    date_cols = [c["name"] for c in columns if c["type"] == "date"]
    if not date_cols:
        return None
    name = date_cols[0]
    dts = [d.replace(tzinfo=None)
           for d in (parse_dt(v) for v in col_values[name]) if d is not None]
    if len(dts) < 2:
        return None
    lo, hi = min(dts), max(dts)
    span_days = (hi - lo).days
    if span_days <= 1:
        unit, key = "hour", lambda d: d.strftime("%Y-%m-%d %H:00")
    elif span_days <= 120:
        unit, key = "day", lambda d: d.date().isoformat()
    elif span_days <= 3600:
        unit, key = "month", lambda d: d.strftime("%Y-%m")
    else:
        unit, key = "year", lambda d: d.strftime("%Y")
    buckets = Counter(key(d) for d in dts)
    # zero-fill days so short ranges read as a timeline
    if unit == "day" and 1 < span_days <= 120:
        from datetime import timedelta
        day = lo.date()
        while day <= hi.date():
            buckets.setdefault(day.isoformat(), 0)
            day += timedelta(days=1)
    return {
        "column": name, "unit": unit,
        "points": [{"label": k, "count": v} for k, v in sorted(buckets.items())],
    }


AGG_CAT_MAX_DISTINCT = 12
AGG_MAX_CAT_COLS = 4
AGG_MAX_NUM_COLS = 6


def build_aggregates(columns, col_values):
    """Deterministic group-by aggregates so downstream consumers (the
    infographic brief, custom sections) never have to compute a number
    themselves. Bounded: only low-cardinality categoricals, only the first
    few numeric columns.

    Rounding conventions: sums/means 2 decimals, shares/rates 1 decimal.
    """
    cat_cols = [c["name"] for c in columns
                if c["type"] == "string"
                and 2 <= c["distinct"] <= AGG_CAT_MAX_DISTINCT][:AGG_MAX_CAT_COLS]
    num_cols = [c["name"] for c in columns
                if c["type"] == "number"][:AGG_MAX_NUM_COLS]
    bool_cols = [c["name"] for c in columns if c["type"] == "boolean"]
    date_cols = [c["name"] for c in columns if c["type"] == "date"]
    if not cat_cols and not date_cols:
        return None

    n_rows = len(next(iter(col_values.values()))) if col_values else 0
    agg = {}

    by_category = {}
    for cat in cat_cols:
        keys = [str(v) if v is not None else None for v in col_values[cat]]
        counts = Counter(k for k in keys if k is not None)
        total = sum(counts.values())
        entry = {
            "counts": dict(counts.most_common()),
            "share_pct": {k: round(100 * c / total, 1)
                          for k, c in counts.most_common()} if total else {},
        }
        sums, means = {}, {}
        for num in num_cols:
            g = {}
            for k, v in zip(keys, col_values[num]):
                x = try_number(v)
                if k is not None and x is not None:
                    g.setdefault(k, []).append(x)
            if g:
                sums[num] = {k: round(sum(xs), 2) for k, xs in g.items()}
                means[num] = {k: round(sum(xs) / len(xs), 2)
                              for k, xs in g.items()}
        rates = {}
        for bcol in bool_cols:
            g = {}
            for k, v in zip(keys, col_values[bcol]):
                b = try_bool(v)
                if k is not None and b is not None:
                    g.setdefault(k, []).append(b)
            if g:
                rates[bcol] = {k: round(100 * sum(bs) / len(bs), 1)
                               for k, bs in g.items()}
        if sums:
            entry["sums"] = sums
            entry["means"] = means
        if rates:
            entry["rates_pct"] = rates
        by_category[cat] = entry
    if by_category:
        agg["by_category"] = by_category

    if date_cols:
        name = date_cols[0]
        parsed = [parse_dt(v) for v in col_values[name]]
        dts = [(i, d.replace(tzinfo=None)) for i, d in enumerate(parsed) if d]
        if len(dts) >= 2:
            lo = min(d for _, d in dts)
            hi = max(d for _, d in dts)
            span_days = (hi - lo).days
            if span_days <= 1:
                unit, key = "hour", lambda d: d.strftime("%Y-%m-%d %H:00")
            elif span_days <= 120:
                unit, key = "day", lambda d: d.date().isoformat()
            elif span_days <= 3600:
                unit, key = "month", lambda d: d.strftime("%Y-%m")
            else:
                unit, key = "year", lambda d: d.strftime("%Y")
            buckets = {}
            for i, d in dts:
                buckets.setdefault(key(d), []).append(i)
            ordered = sorted(buckets.items())
            counts = {k: len(idx) for k, idx in ordered}
            time_entry = {"column": name, "unit": unit, "counts": counts}
            sums = {}
            for num in num_cols:
                vals = col_values[num]
                s = {k: round(sum(x for x in (try_number(vals[i]) for i in idx)
                                  if x is not None), 2)
                     for k, idx in ordered}
                sums[num] = s
            if sums:
                time_entry["sums"] = sums
            if len(counts) >= 2:
                peak_bucket = max(counts, key=lambda k: counts[k])
                labels = [k for k, _ in ordered]
                pos = labels.index(peak_bucket)
                peak = {"bucket": peak_bucket, "count": counts[peak_bucket]}
                if pos > 0:
                    prev = counts[labels[pos - 1]]
                    if prev:
                        peak["pct_vs_prev"] = round(
                            100 * (counts[peak_bucket] - prev) / prev, 1)
                others = [c for k, c in counts.items() if k != peak_bucket]
                mean_others = sum(others) / len(others) if others else 0
                if mean_others:
                    peak["pct_vs_mean_others"] = round(
                        100 * (counts[peak_bucket] - mean_others) / mean_others, 1)
                time_entry["peak"] = peak
            agg["by_time"] = time_entry

    agg["row_basis"] = n_rows
    return agg or None


def normalize_table(headers, rows, label, source_format, max_rows, note=None):
    """headers: list[str]; rows: list[list] aligned with headers."""
    headers = [str(h) if clean(h) is not None else f"col_{i+1}"
               for i, h in enumerate(headers)]
    # Duplicate names would collapse into one column-major bucket and corrupt
    # every downstream stat — keep each physical column its own logical column.
    seen = Counter()
    uniq = []
    for h in headers:
        seen[h] += 1
        uniq.append(h if seen[h] == 1 else f"{h} ({seen[h]})")
    headers = uniq
    if len(headers) > MAX_COLS:
        note = ((note + " · ") if note else "") + \
            f"showing first {MAX_COLS} of {len(headers)} columns"
        headers = headers[:MAX_COLS]
        rows = [r[:MAX_COLS] for r in rows]

    # column-major values (padded)
    col_values = {h: [] for h in headers}
    for r in rows:
        for i, h in enumerate(headers):
            col_values[h].append(clean(r[i]) if i < len(r) else None)

    columns = [profile_column(h, col_values[h]) for h in headers]
    timeseries = build_timeseries(columns, col_values)
    aggregates = build_aggregates(columns, col_values)

    total_rows = len(rows)
    null_cells = sum(c["nulls"] for c in columns)
    cell_count = total_rows * len(headers)
    type_tally = Counter(c["type"] for c in columns)

    capped = total_rows > max_rows
    display_rows = [[cell_repr(c) for c in r[:len(headers)]]
                    for r in rows[:max_rows]]
    if capped:
        note = ((note + " · ") if note else "") + \
            f"table shows the first {max_rows:,} of {total_rows:,} rows (stats cover all rows)"

    return {
        "type": "table",
        "label": label,
        "stats": {
            "row_count": total_rows,
            "col_count": len(headers),
            "cell_count": cell_count,
            "null_cells": null_cells,
            "filled_pct": round(100 * (1 - null_cells / cell_count), 1) if cell_count else 0,
            "numeric_cols": type_tally.get("number", 0),
            "text_cols": type_tally.get("string", 0),
            "date_cols": type_tally.get("date", 0),
            "bool_cols": type_tally.get("boolean", 0),
            "source_format": source_format,
        },
        "columns": columns,
        "timeseries": timeseries,
        "aggregates": aggregates,
        "rows": display_rows,
        "capped": capped,
        "note": note,
    }


def records_to_table(records, label, source_format, max_rows, note=None):
    """List of dicts → table (union of keys, first-seen order)."""
    headers, seen = [], set()
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    rows = [[rec.get(h) for h in headers] for rec in records]
    return normalize_table(headers, rows, label, source_format, max_rows, note)


def looks_like_records(data):
    """A JSON list that deserves the tabular dashboard: mostly flat dicts
    sharing most of their keys."""
    if not isinstance(data, list) or len(data) < 2:
        return False
    sample = data[:200]
    dicts = [d for d in sample if isinstance(d, dict)]
    if len(dicts) < len(sample) * 0.95:
        return False
    key_sets = [set(d.keys()) for d in dicts]
    all_keys = set().union(*key_sets)
    if not all_keys or len(all_keys) > MAX_COLS * 2:
        return False
    common = [k for k in all_keys
              if sum(1 for ks in key_sets if k in ks) >= len(dicts) * 0.6]
    if len(common) < max(1, len(all_keys) * 0.4):
        return False
    # mostly scalar values?
    scalar = nested = 0
    for d in dicts[:50]:
        for v in d.values():
            if isinstance(v, (dict, list)):
                nested += 1
            else:
                scalar += 1
    return scalar >= nested * 2


# ---------------------------------------------------------------- text dataset

_HEADING_RE = re.compile(r"^(#{1,4})\s+(.+?)\s*#*\s*$", re.M)
_STOPWORDS = set("""a an the and or but if then else for while of to in on at by
with from as is are was were be been being it its this that these those i you he
she we they them his her their our your my me us not no yes do does did done can
could should would will shall may might must have has had having so than too very
just about into over under out up down off again further once here there when
where why how all any both each few more most other some such only own same s t
don now what which who whom""".split())


def normalize_text(content, label, fmt, max_chars):
    lines = content.split("\n")
    words = word_count(content)
    headings = [{"level": len(m.group(1)), "text": m.group(2).strip()}
                for m in _HEADING_RE.finditer(content)] if fmt == "markdown" else []
    code_blocks = content.count("```") // 2 if fmt == "markdown" else 0
    links = len(re.findall(r"\[[^\]]+\]\([^)]+\)", content)) \
        + len(re.findall(r"(?<!\()https?://\S+", content))
    paragraphs = len([p for p in re.split(r"\n\s*\n", content) if p.strip()])

    tokens = re.findall(r"[A-Za-z][A-Za-z'’-]{2,}", content.lower())
    counter = Counter(t for t in tokens if t not in _STOPWORDS)
    top_words = [{"word": w, "count": c} for w, c in counter.most_common(12)]

    shown, capped = cap_text(content, max_chars)
    return {
        "type": "text",
        "label": label,
        "format": fmt,
        "stats": {
            "words": words,
            "chars": len(content),
            "lines": len(lines),
            "paragraphs": paragraphs,
            "headings": len(headings),
            "code_blocks": code_blocks,
            "links": links,
            "reading_minutes": max(1, round(words / 220)) if words else 0,
        },
        "outline": headings[:60],
        "top_words": top_words,
        "content": shown,
        "capped": capped,
        "note": f"document truncated to {max_chars:,} chars for display" if capped else None,
    }


# ---------------------------------------------------------------- log dataset

_LEVEL_RE = re.compile(
    r"\b(TRACE|DEBUG|INFO|NOTICE|WARN(?:ING)?|ERROR|SEVERE|CRITICAL|FATAL|PANIC)\b",
    re.I)
_TS_RE = re.compile(
    r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?"      # ISO-ish
    r"|\d{2}/[A-Za-z]{3}/\d{4}:\d{2}:\d{2}:\d{2}"     # Apache CLF
    r"|[A-Za-z]{3}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}")    # syslog
_SOURCE_RE = re.compile(r"\[([A-Za-z0-9_.\-/:]{2,40})\]")

_LEVEL_MAP = {"WARNING": "WARN", "SEVERE": "ERROR", "CRITICAL": "ERROR",
              "FATAL": "ERROR", "PANIC": "ERROR", "NOTICE": "INFO",
              "TRACE": "DEBUG"}


def _log_ts(match_text):
    txt = match_text.replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%d/%b/%Y:%H:%M:%S", "%b %d %H:%M:%S"):
        try:
            dt = datetime.strptime(txt, fmt)
            if dt.year == 1900:  # syslog has no year
                dt = dt.replace(year=datetime.now().year)
            return dt
        except ValueError:
            continue
    return None


def looks_like_log(sample_lines):
    lines = [l for l in sample_lines if l.strip()]
    if len(lines) < 5:
        return False
    hits = sum(1 for l in lines
               if _TS_RE.search(l[:64]) or _LEVEL_RE.search(l[:96]))
    return hits >= len(lines) * 0.5


def normalize_log(content, label, max_lines):
    raw_lines = content.split("\n")
    if raw_lines and raw_lines[-1] == "":
        raw_lines.pop()

    levels = Counter()
    by_date = Counter()
    by_hour = [0] * 24
    sources = Counter()
    all_ts = []
    entries = []

    for raw in raw_lines:
        lvl = None
        m = _LEVEL_RE.search(raw[:120])
        if m:
            lvl = m.group(1).upper()
            lvl = _LEVEL_MAP.get(lvl, lvl)
            levels[lvl] += 1
        ts = None
        tm = _TS_RE.search(raw[:64])
        if tm:
            ts = _log_ts(tm.group(0))
            if ts:
                by_date[ts.date().isoformat()] += 1
                by_hour[ts.hour] += 1
                all_ts.append(ts)
        sm = _SOURCE_RE.search(raw[:160])
        if sm and not _TS_RE.match(sm.group(1)):
            sources[sm.group(1)] += 1
        entries.append({
            "raw": raw if len(raw) <= 600 else raw[:600] + "…",
            "level": lvl,
        })

    total = len(raw_lines)
    date_range, days_span = None, 0
    if all_ts:
        lo, hi = min(all_ts), max(all_ts)
        date_range = [lo.isoformat(), hi.isoformat()]
        days_span = (hi.date() - lo.date()).days + 1
        if 1 < days_span <= 400:
            from datetime import timedelta
            day = lo.date()
            while day <= hi.date():
                by_date.setdefault(day.isoformat(), 0)
                day += timedelta(days=1)

    err = levels.get("ERROR", 0)
    capped = total > max_lines
    note = (f"viewer shows the first {max_lines:,} of {total:,} lines "
            f"(stats cover all lines)") if capped else None

    return {
        "type": "log",
        "label": label,
        "stats": {
            "total_lines": total,
            "leveled_lines": sum(levels.values()),
            "timestamped_lines": len(all_ts),
            "levels": [{"level": k, "count": v}
                       for k, v in levels.most_common()],
            "error_count": err,
            "error_pct": round(100 * err / total, 1) if total else 0,
            "by_date": [{"date": d, "count": c}
                        for d, c in sorted(by_date.items())],
            "by_hour": by_hour,
            "date_range": date_range,
            "days_span": days_span,
            "top_sources": [{"name": n, "count": c}
                            for n, c in sources.most_common(10)],
        },
        "lines": entries[:max_lines],
        "capped": capped,
        "note": note,
    }


# ---------------------------------------------------------------- generic (JSON tree)

def walk_stats(node, depth, acc):
    acc["total"] += 1
    acc["max_depth"] = max(acc["max_depth"], depth)
    if isinstance(node, dict):
        acc["types"]["object"] += 1
        for k, v in node.items():
            acc["keys"][k] += 1
            walk_stats(v, depth + 1, acc)
    elif isinstance(node, list):
        acc["types"]["array"] += 1
        for v in node:
            walk_stats(v, depth + 1, acc)
    elif isinstance(node, str):
        acc["types"]["string"] += 1
    elif isinstance(node, bool):
        acc["types"]["boolean"] += 1
    elif isinstance(node, (int, float)):
        acc["types"]["number"] += 1
    else:
        acc["types"]["null"] += 1


def cap_arrays(node, cap):
    if isinstance(node, dict):
        return {k: cap_arrays(v, cap) for k, v in node.items()}
    if isinstance(node, list):
        return [cap_arrays(v, cap) for v in node[:cap]]
    return node


def normalize_generic(data, label):
    acc = {"total": 0, "max_depth": 0, "types": Counter(), "keys": Counter()}
    walk_stats(data, 0, acc)

    payload, capped, note = data, False, None
    if len(json.dumps(data, ensure_ascii=False)) > MAX_GENERIC_BYTES:
        payload = cap_arrays(data, GENERIC_ARRAY_CAP)
        capped = True
        note = (f"Large file: arrays truncated to their first "
                f"{GENERIC_ARRAY_CAP} items for display.")

    return {
        "type": "generic",
        "label": label,
        "stats": {
            "root_type": "array" if isinstance(data, list) else
                         "object" if isinstance(data, dict) else "value",
            "total_nodes": acc["total"],
            "max_depth": acc["max_depth"],
            "type_counts": dict(acc["types"]),
            "top_keys": [
                {"key": k, "count": c} for k, c in acc["keys"].most_common(15)
            ],
        },
        "data": payload,
        "capped": capped,
        "note": note,
    }


# ---------------------------------------------------------------- loaders

def load_csv(path, delimiter, label, max_rows):
    with open(path, encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(64 * 1024)
        fh.seek(0)
        if delimiter is None:
            try:
                delimiter = csv.Sniffer().sniff(sample, ",;\t|").delimiter
            except csv.Error:
                delimiter = ","
        reader = csv.reader(fh, delimiter=delimiter)
        rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [normalize_generic([], label)]
    headers, body = rows[0], rows[1:]
    if not body:  # single row — treat first row as data with generated headers
        headers = [f"col_{i+1}" for i in range(len(rows[0]))]
        body = rows
    fmt = "tsv" if delimiter == "\t" else "csv"
    return [normalize_table(headers, body, label, fmt, max_rows)]


def load_jsonl(path, label, max_rows):
    records, bad = [], 0
    with open(path, encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                bad += 1
    note = f"{bad} unparseable line(s) skipped" if bad else None
    if records and all(isinstance(r, dict) for r in records[:200]):
        return [records_to_table(records, label, "jsonl", max_rows, note)]
    return [normalize_generic(records, label)]


def load_json(path, label, max_rows):
    with open(path, encoding="utf-8-sig") as fh:
        data = json.load(fh)
    if looks_like_records(data):
        return [records_to_table(data, label, "json", max_rows)]
    return [normalize_generic(data, label)]


def load_xlsx(path, label, max_rows):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "Excel input needs openpyxl — install with:\n"
            "  pip install openpyxl --break-system-packages")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    datasets = []
    sheets = wb.sheetnames[:SHEET_CAP]
    for sheet in sheets:
        ws = wb[sheet]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(clean(c) is not None for c in r)]
        if not rows:
            continue
        headers, body = rows[0], rows[1:]
        if not body:
            headers = [f"col_{i+1}" for i in range(len(rows[0]))]
            body = rows
        lbl = label if len(sheets) == 1 else f"{label} · {sheet}"
        datasets.append(normalize_table(headers, body, lbl, "xlsx", max_rows))
    wb.close()
    if len(wb.sheetnames) > SHEET_CAP and datasets:
        datasets[0]["note"] = ((datasets[0]["note"] + " · ") if datasets[0]["note"] else "") + \
            f"first {SHEET_CAP} of {len(wb.sheetnames)} sheets shown"
    return datasets or [normalize_generic({}, label)]


def load_sqlite(path, label, max_rows):
    import sqlite3
    con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    try:
        cur = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%' ORDER BY name")
        tables = [r[0] for r in cur.fetchall()]
        datasets = []
        for t in tables[:SQL_TABLE_CAP]:
            cur = con.execute(f'SELECT * FROM "{t}" LIMIT {SQL_ROW_SCAN_CAP}')
            headers = [d[0] for d in cur.description]
            body = [list(r) for r in cur.fetchall()]
            lbl = label if len(tables) == 1 else f"{label} · {t}"
            note = (f"first {SQL_ROW_SCAN_CAP:,} rows scanned"
                    if len(body) == SQL_ROW_SCAN_CAP else None)
            datasets.append(
                normalize_table(headers, body, lbl, "sqlite", max_rows, note))
        if len(tables) > SQL_TABLE_CAP and datasets:
            datasets[0]["note"] = ((datasets[0]["note"] + " · ") if datasets[0]["note"] else "") + \
                f"first {SQL_TABLE_CAP} of {len(tables)} tables shown"
        return datasets or [normalize_generic({"tables": tables}, label)]
    finally:
        con.close()


def load_text_like(path, label, max_text_chars, max_log_lines, force=None):
    with open(path, encoding="utf-8", errors="replace") as fh:
        content = fh.read()
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if force == "log" or (force is None and (
            ext == "log" or looks_like_log(content.split("\n")[:200]))):
        return [normalize_log(content, label, max_log_lines)]
    fmt = "markdown" if ext in ("md", "markdown") or (
        force is None and _HEADING_RE.search(content[:4000])) else "plain"
    return [normalize_text(content, label, fmt, max_text_chars)]


# ---------------------------------------------------------------- driver

def normalize_file(path, opts):
    """Returns a list of dataset dicts (multi-sheet/table files yield several)."""
    base = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    max_rows = opts["max_rows"]

    if ext == "csv":
        return load_csv(path, ",", base, max_rows)
    if ext == "tsv":
        return load_csv(path, "\t", base, max_rows)
    if ext in ("jsonl", "ndjson"):
        return load_jsonl(path, base, max_rows)
    if ext == "json":
        return load_json(path, base, max_rows)
    if ext in ("xlsx", "xlsm"):
        return load_xlsx(path, base, max_rows)
    if ext in ("sqlite", "sqlite3", "db"):
        return load_sqlite(path, base, max_rows)
    if ext in ("md", "markdown"):
        return load_text_like(path, base, opts["max_text_chars"],
                              opts["max_log_lines"], force="text")
    if ext == "log":
        return load_text_like(path, base, opts["max_text_chars"],
                              opts["max_log_lines"], force="log")

    # unknown extension — sniff content
    with open(path, "rb") as fh:
        head = fh.read(8192)
    if head[:4] == b"SQLite"[:4] and head[:16].startswith(b"SQLite format 3"):
        return load_sqlite(path, base, max_rows)
    try:
        text_head = head.decode("utf-8")
    except UnicodeDecodeError:
        raise SystemExit(f"{base}: binary file with unrecognized format — "
                         "supported: csv/tsv/json/jsonl/xlsx/sqlite/md/txt/log")
    stripped = text_head.lstrip()
    if stripped[:1] in ("{", "["):
        try:
            return load_json(path, base, max_rows)
        except json.JSONDecodeError:
            pass
    first = stripped.split("\n", 1)[0] if stripped else ""
    if first.count(",") >= 2 or first.count("\t") >= 2:
        try:
            return load_csv(path, None, base, max_rows)
        except Exception:
            pass
    return load_text_like(path, base, opts["max_text_chars"],
                          opts["max_log_lines"])


def build_model(paths, title, subtitle, accent, **kw):
    opts = {
        "max_rows": kw.get("max_rows", DEFAULT_MAX_ROWS),
        "max_text_chars": kw.get("max_text_chars", DEFAULT_MAX_TEXT_CHARS),
        "max_log_lines": kw.get("max_log_lines", DEFAULT_MAX_LOG_LINES),
    }
    datasets = []
    for path in paths:
        for ds in normalize_file(path, opts):
            ds["id"] = f"ds{len(datasets)}"
            datasets.append(ds)
    return {
        "meta": {
            "title": title,
            "subtitle": subtitle,
            "accent": accent,
            "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "sources": [os.path.basename(p) for p in paths],
        },
        "datasets": datasets,
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("inputs", nargs="+", help="data file(s) to normalize")
    ap.add_argument("-o", "--output", default="normalized.json")
    ap.add_argument("--title", default="Data Explorer")
    ap.add_argument("--subtitle", default="")
    ap.add_argument("--accent", default="#7c6cff")
    ap.add_argument("--max-rows", type=int, default=DEFAULT_MAX_ROWS)
    ap.add_argument("--max-text-chars", type=int, default=DEFAULT_MAX_TEXT_CHARS)
    ap.add_argument("--max-log-lines", type=int, default=DEFAULT_MAX_LOG_LINES)
    args = ap.parse_args(argv)

    model = build_model(args.inputs, args.title, args.subtitle, args.accent,
                        max_rows=args.max_rows,
                        max_text_chars=args.max_text_chars,
                        max_log_lines=args.max_log_lines)
    with open(args.output, "w", encoding="utf-8") as fh:
        json.dump(model, fh, ensure_ascii=False)
    types = ", ".join(d["type"] for d in model["datasets"])
    print(f"Wrote {args.output} — {len(model['datasets'])} dataset(s): {types}")
    return model


if __name__ == "__main__":
    sys.exit(0 if main() else 1)
